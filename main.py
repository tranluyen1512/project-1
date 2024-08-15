from kivy.uix.textinput import TextInput
from kivy.app import App 
from kivy.uix.label import Label
from kivy.uix.image import Image
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from openpyxl import Workbook, load_workbook

class MyApp(App):
    def build(self):
        layout = BoxLayout(orientation="vertical", padding = {80,40,80,80})
        self.icon = "hinh1.png"
        self.title = "App tính tiền trọ Ngọc Bích"
        main_label = Label(
            text = "App tính tiền trọ Ngọc Bích",
            font_size=30, bold=True,color=(1, 0, 0, 1) 
        )
        layout.add_widget(main_label)


        ngay_label = BoxLayout(orientation='horizontal')
        layout.add_widget(ngay_label)

  # ngày tháng năm
        label = Label(text = "Ngày :",pos_hint={"center_x":0.1})
        ngay_label.add_widget(label)
        self.ngay_input = TextInput(hint_text="Nhập ngày tháng năm:", size_hint_y=0.7)
        ngay_label.add_widget(self.ngay_input)


# 1
        image_label_box = BoxLayout(orientation="horizontal")
        layout.add_widget(image_label_box)

   
        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)
        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)
        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)
        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)
        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)
        image_box = BoxLayout(orientation="vertical")  
        image_label_box.add_widget(image_box)

        img1 = Image(source="hinh1.png", size_hint=(0.6, 0.6))
        image_box.add_widget(img1)

        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)

        label1 = Label(
            text="ĐIỆN", halign="center", valign="middle", size_hint_y=0.7  
        )
        label_box.add_widget(label1)



        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)
        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)
        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)
        image_box = BoxLayout(orientation="vertical")  
        image_label_box.add_widget(image_box)

        img2 = Image(source="hinh2.jpg", size_hint=(0.6, 0.6))
        image_box.add_widget(img2)

        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)

        label2 = Label(
            text="NƯỚC", halign="center", valign="middle", size_hint_y=0.2 
        )
        label_box.add_widget(label2)

        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)

       #2 


        image_label_box = BoxLayout(orientation="horizontal", spacing=0)
        layout.add_widget(image_label_box)

        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)

        label1 = Label(
            text="MỚI", halign="center", valign="middle", size_hint_y=0.7  
        )
        label_box.add_widget(label1)
        self.dienmoi_input = TextInput(hint_text="Nhập số điện mới:", size_hint_y=0.7)
        image_label_box.add_widget(self.dienmoi_input)
        self.nuocmoi_input = TextInput(hint_text="Nhập số nước mới:", size_hint_y=0.7)
        image_label_box.add_widget(self.nuocmoi_input)


        image_label_box = BoxLayout(orientation="horizontal", spacing=0)
        layout.add_widget(image_label_box)
        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)

        label2 = Label(
            text="CŨ", halign="center", valign="middle", size_hint_y=0.2 
        )
        label_box.add_widget(label2)

        self.diencu_input = TextInput(hint_text="Nhập số điện cũ:", size_hint_y=0.7)
        image_label_box.add_widget(self.diencu_input)
        self.nuoccu_input = TextInput(hint_text="Nhập số nước cũ:", size_hint_y=0.7)
        image_label_box.add_widget(self.nuoccu_input)

        image_label_box = BoxLayout(orientation="horizontal", spacing=0)
        layout.add_widget(image_label_box)




        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)

        calc_button = Button(text="KẾT QUẢ",color=(1, 0, 0, 1) , bold= True ,size_hint=(0.5, 0.5), pos_hint={"center_x": 0.5})
        calc_button.bind(on_press=self.calculate_bill)
        label_box.add_widget(calc_button)

        self.tongcondiencu_input = TextInput(hint_text="Số điện mới -số điện cũ",readonly = True, size_hint_y=0.7)
        image_label_box.add_widget(self.tongcondiencu_input)
        self.tongconnuoccu_input = TextInput(hint_text="số nước mới - số nước cũ", readonly = True,size_hint_y=0.7)
        image_label_box.add_widget(self.tongconnuoccu_input)
    


# 2.2
        image_label_box = BoxLayout(orientation="horizontal", spacing=0)
        layout.add_widget(image_label_box)

        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)
        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)

        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)

        label1 = Label(
            text="SỬ DỤNG", halign="center", valign="middle", size_hint_y=0.7  
        )
        label_box.add_widget(label1)

        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)

        label2 = Label(
            text="GIÁ", halign="center", valign="middle", size_hint_y=0.2 
        )
        label_box.add_widget(label2)
        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)

        label3 = Label(
            text="TỔNG", halign="center", valign="middle", size_hint_y=0.2 
        )
        label_box.add_widget(label3)

#3

        image_label_box = BoxLayout(orientation="horizontal", spacing=0)
        layout.add_widget(image_label_box)

   
        image_box = BoxLayout(orientation="vertical")  
        image_label_box.add_widget(image_box)

        img1 = Image(source="hinh3.png", size_hint=(0.3, 0.3), pos_hint={"center_x": 0.3,"center_y":0.7})
        image_box.add_widget(img1)

        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)

        label1 = Label(
            text="TIỀN PHÒNG", halign="left", valign="middle", size_hint_y=0.7  
        )
        label_box.add_widget(label1)
        self.tenphong_input = TextInput(hint_text="Nhập tên phòng:", size_hint_y=0.7)
        image_label_box.add_widget(self.tenphong_input)




        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)
        label1 = Label(
            text="GIÁ PHÒNG", valign="middle", size_hint_y=0.7  
        )
        label_box.add_widget(label1)
        self.giaphong_input = TextInput(hint_text="", size_hint_y=0.7)
        image_label_box.add_widget(self.giaphong_input)



#4 

        image_label_box = BoxLayout(orientation="horizontal", spacing=0)
        layout.add_widget(image_label_box)

   
        image_box = BoxLayout(orientation="vertical")  
        image_label_box.add_widget(image_box)

        img1 = Image(source="hinh1.png", size_hint=(0.3, 0.3), pos_hint={"center_x": 0.3,"center_y":0.7})
        image_box.add_widget(img1)

        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)

        label1 = Label(
            text="TIỀN ĐIỆN", halign="center", valign="middle", size_hint_y=0.7  
        )
        label_box.add_widget(label1)
        self.tiendien_input = TextInput(hint_text="Nhập số điện:", size_hint_y=0.7)
        image_label_box.add_widget(self.tiendien_input)
        self.giadien_input = TextInput(hint_text="5.000 VNĐ", readonly=True, size_hint_y=0.7)
        image_label_box.add_widget(self.giadien_input)
        self.tongdien_input = TextInput(hint_text="", size_hint_y=0.7, readonly = True, )
        image_label_box.add_widget(self.tongdien_input)





        #5
        image_label_box = BoxLayout(orientation="horizontal", spacing=0)
        layout.add_widget(image_label_box)

   
        image_box = BoxLayout(orientation="vertical")  
        image_label_box.add_widget(image_box)

        img1 = Image(source="hinh2.jpg", size_hint=(0.3, 0.3), pos_hint={"center_x": 0.3,"center_y":0.7})
        image_box.add_widget(img1)

        label_box = BoxLayout(orientation="vertical")
        image_label_box.add_widget(label_box)

        label1 = Label(
            text="TIỀN NƯỚC", halign="center", valign="middle", size_hint_y=0.7  
        )
        label_box.add_widget(label1)
        self.tiennuoc_input = TextInput(hint_text="Nhập số nước:", size_hint_y=0.7)
        image_label_box.add_widget(self.tiennuoc_input)
        self.gianuoc_input = TextInput(hint_text="15.000 VNĐ", readonly=True,size_hint_y=0.7)
        image_label_box.add_widget(self.gianuoc_input)
        self.tongnuoc_input = TextInput(hint_text="",readonly = True, size_hint_y=0.7)
        image_label_box.add_widget(self.tongnuoc_input)





# 6
        image_label_box = BoxLayout(orientation="horizontal", spacing=10)
        layout.add_widget(image_label_box)


        image_box = BoxLayout(orientation="vertical")  
        image_label_box.add_widget(image_box)
        export_button = Button(text="Xuất ra Excel", color=(0, 1, 0, 1), bold=True, size_hint=(1, 0.6))
        export_button.bind(on_press=self.export_to_excel)
        image_label_box.add_widget(export_button)




        label_box = BoxLayout(orientation="horizontal")
        image_label_box.add_widget(label_box)
        calc_button = Button(text="Thành tiền",color=(0, 1, 0, 1) , bold= True ,size_hint=(1, 0.6))
        calc_button.bind(on_press=self.calculate_bill)
        label_box.add_widget(calc_button)

        # Nhãn hiển thị kết quả
        self.thanhtien_input = TextInput(hint_text="",readonly = True, size_hint_y=0.7)
        image_label_box.add_widget(self.thanhtien_input)


        return layout
    


    def calculate_bill(self, instance):
        try:

            dien_moi = int(self.dienmoi_input.text)
            dien_cu = int(self.diencu_input.text)
            nuoc_moi = int(self.nuocmoi_input.text)
            nuoc_cu = int(self.nuoccu_input.text)


            dien_tieu_thu = dien_moi - dien_cu
            nuoc_tieu_thu = nuoc_moi - nuoc_cu

            # Logic tính tiền điện, nước ở đây (bạn thay thế bằng công thức tính toán thực tế)
            gia_dien = 5000  # Ví dụ: giá điện 5000 đồng/kWh
            gia_nuoc = 15000  # Ví dụ: giá nước 15000 đồng/m3
            self.tongcondiencu_input.text = f"Tổng điện tiêu thụ: {dien_tieu_thu:} kWh"
            self.tongconnuoccu_input.text = f"Tổng nước tiêu thụ: {nuoc_tieu_thu:} m³"
            a = int(dien_tieu_thu)
            b = int(nuoc_tieu_thu)
            c = int (dien_tieu_thu) * int(gia_dien)
            d = int(nuoc_tieu_thu)* int(gia_nuoc)
        
            self.tongdien_input.text = f"{c:,} VNĐ"
            self.tongnuoc_input.text = f"{d:,} VNĐ"
            self.tiendien_input.text = f"{a} kWh"
            self.tiennuoc_input.text = f"{b} m³"
            gia_phong = int(self.giaphong_input.text)


            tong_tien = dien_tieu_thu * gia_dien + nuoc_tieu_thu * gia_nuoc + gia_phong
            self.thanhtien_input.text= f" {tong_tien:,} VNĐ"
        except ValueError:
            self.thanhtien_input.text = "nhập tiền phòng!"



    def export_to_excel(self, instance):
        workbook = Workbook()
        sheet = workbook.active
        row = 2
        # Viết tiêu đề cột
        sheet['A1'] = 'Ngày'
        sheet['B1'] = 'Họ tên phòng'
        sheet['C1'] = 'Số điện mới'
        sheet['D1'] = 'Số điện cũ'
        sheet['E1'] = 'Số điện tiêu thụ'
        sheet['F1'] = 'Số nước mới'
        sheet['G1'] = 'Số nước cũ'
        sheet['H1'] = 'Số nước tiêu thụ'
        sheet['I1'] = 'Giá phòng'
        sheet['J1'] = 'Thành Tiền'

        # Viết dữ liệu vào các hàng
        
        
        sheet.cell(row=row, column=1, value=self.ngay_input.text)
        sheet.cell(row=row, column=2, value=self.tenphong_input.text)
        sheet.cell(row=row, column=3, value=self.dienmoi_input.text)
        sheet.cell(row=row, column=4, value=self.diencu_input.text)
        sheet.cell(row=row, column=5, value=self.tongcondiencu_input.text)
        sheet.cell(row=row, column=6, value=self.nuocmoi_input.text)
        sheet.cell(row=row, column=7, value=self.nuoccu_input.text)
        sheet.cell(row=row, column=8, value=self.tongconnuoccu_input.text)
        sheet.cell(row=row, column=9, value=self.giaphong_input.text)
        sheet.cell(row=row, column=10, value=self.thanhtien_input.text)
        workbook.save('bill.xlsx')
        row +=1
        try:
            # Nếu file đã tồn tại, mở và load dữ liệu cũ
            existing_workbook = load_workbook('bill.xlsx')
            sheet = existing_workbook.active
            # Tìm dòng cuối cùng có dữ liệu để tiếp tục ghi
            row = row + 1
        except FileNotFoundError:
            pass  # Nếu file chưa tồn tại thì tạo mới

        # Lưu file Excel
        workbook.save('bill.xlsx')
        



     #   def calculate_bill(self, instance):
     #       try:
        # Kiểm tra dữ liệu đầu vào
     #           dien_moi = int(self.dienmoi_input.text)
     #           if dien_moi < 0:
     #               raise ValueError("Số điện mới phải lớn hơn 0")

        # Các kiểm tra tương tự cho các trường khác

        # ... (phần tính toán)

     #       except ValueError as e:
      #          self.thanhtien_input.text = str(e)




#run app
if __name__ == "__main__":
    app =MyApp()
    app.run()